#!/usr/bin/env python3
"""
Skapa Azure DevOps work items från Freshservice-ärendelistan (Excel).

Källa:  "Ärenden Freshservice.xlsx" (kolumner: ID, Kategori, Typ,
        Prioritering av verksamhet, Rubrik, Kort beskrivning, Status, Kommentar Alex)
Mål:    https://dev.azure.com/BDO-Sweden/Freshservice

Användning:
    export AZDO_PAT=<din-personal-access-token>        # behörighet: Work Items (Read & Write)
    python3 scripts/create_freshservice_tickets.py "/path/till/Ärenden Freshservice.xlsx" --dry-run
    python3 scripts/create_freshservice_tickets.py "/path/till/Ärenden Freshservice.xlsx"

Beroenden:
    pip3 install openpyxl requests
"""

import argparse
import base64
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Saknar openpyxl. Kör: pip3 install openpyxl")

try:
    import requests
except ImportError:
    sys.exit("Saknar requests. Kör: pip3 install requests")


# ---------------------------------------------------------------------------
# Konfiguration
# ---------------------------------------------------------------------------
ORG = "BDO-Sweden"
PROJECT = "Freshservice"
API_VERSION = "7.1"

# Freshservice "Typ" -> Azure DevOps work item-typ.
# Bug och Task finns i alla processmallar (Basic/Agile/Scrum/CMMI).
TYPE_MAP = {
    "bug": "Bug",
    "bug / krav ?": "Bug",
    "bug / designgap": "Bug",
}
DEFAULT_WORK_ITEM_TYPE = "Task"  # Förbättring, Konfiguration, Begränsning m.fl.

# "Prioritering av verksamhet" -> Microsoft.VSTS.Common.Priority (1 = högst)
PRIORITY_MAP = {
    "hög": 1,
    "hog": 1,
    "medel": 2,
    "låg": 3,
    "lag": 3,
}


def map_type(typ: str) -> str:
    return TYPE_MAP.get((typ or "").strip().lower(), DEFAULT_WORK_ITEM_TYPE)


def map_priority(prio: str):
    return PRIORITY_MAP.get((prio or "").strip().lower())


def html_escape(text: str) -> str:
    return (
        (text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def read_rows(path: str):
    """Läs Excel och returnera en lista av dict per ärende."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Hitta rubrikraden (den som innehåller 'ID' i kolumn A).
    header_row = None
    headers = []
    for r in range(1, ws.max_row + 1):
        first = ws.cell(row=r, column=1).value
        if isinstance(first, str) and first.strip().upper() == "ID":
            header_row = r
            headers = [
                (ws.cell(row=r, column=c).value or "").strip()
                for c in range(1, ws.max_column + 1)
            ]
            break
    if header_row is None:
        sys.exit("Hittade ingen rubrikrad med 'ID' i första kolumnen.")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        record = {headers[i]: values[i] for i in range(len(headers))}
        if not (record.get("ID") or record.get("Rubrik")):
            continue  # hoppa över tomma rader
        rows.append(record)
    return rows


def build_description(rec: dict) -> str:
    """Bygg en HTML-beskrivning med all kontext från Excel-raden."""
    parts = []
    desc = rec.get("Kort beskrivning")
    if desc:
        parts.append(f"<p>{html_escape(str(desc))}</p>")

    meta = []
    for label in ("ID", "Kategori", "Typ", "Prioritering av verksamhet", "Status"):
        val = rec.get(label)
        if val:
            meta.append(f"<li><b>{html_escape(label)}:</b> {html_escape(str(val))}</li>")
    komm = rec.get("Kommentar Alex")
    if komm:
        meta.append(f"<li><b>Kommentar:</b> {html_escape(str(komm))}</li>")
    if meta:
        parts.append("<p><b>Freshservice-metadata</b></p><ul>" + "".join(meta) + "</ul>")
    return "".join(parts)


def build_patch(rec: dict) -> list:
    """Bygg JSON Patch-dokument för work item-skapande."""
    title = str(rec.get("Rubrik") or rec.get("ID") or "Utan rubrik").strip()
    ops = [
        {"op": "add", "path": "/fields/System.Title", "value": title},
        {"op": "add", "path": "/fields/System.Description", "value": build_description(rec)},
    ]

    prio = map_priority(rec.get("Prioritering av verksamhet"))
    if prio:
        ops.append({"op": "add", "path": "/fields/Microsoft.VSTS.Common.Priority", "value": prio})

    # Taggar: FS-ID + kategori + verksamhetsprioritet (för spårbarhet/filtrering).
    tags = []
    if rec.get("ID"):
        tags.append(str(rec["ID"]).strip())
    if rec.get("Kategori"):
        tags.append(str(rec["Kategori"]).strip())
    if rec.get("Prioritering av verksamhet"):
        tags.append(f"Prio: {str(rec['Prioritering av verksamhet']).strip()}")
    tags.append("Freshservice-import")
    if tags:
        ops.append({"op": "add", "path": "/fields/System.Tags", "value": "; ".join(tags)})

    return ops


def create_work_item(session, wi_type: str, patch: list):
    url = (
        f"https://dev.azure.com/{ORG}/{PROJECT}/_apis/wit/workitems/"
        f"${wi_type}?api-version={API_VERSION}"
    )
    resp = session.post(
        url,
        json=patch,
        headers={"Content-Type": "application/json-patch+json"},
    )
    resp.raise_for_status()
    return resp.json()


def main():
    global ORG, PROJECT
    parser = argparse.ArgumentParser(description="Skapa ADO work items från Freshservice-Excel.")
    parser.add_argument("excel", help="Sökväg till Excel-filen")
    parser.add_argument("--dry-run", action="store_true", help="Visa vad som skulle skapas utan att anropa API.")
    parser.add_argument("--org", default=ORG)
    parser.add_argument("--project", default=PROJECT)
    args = parser.parse_args()

    ORG, PROJECT = args.org, args.project

    rows = read_rows(args.excel)
    print(f"Läste {len(rows)} ärenden från {args.excel}\n")

    session = None
    if not args.dry_run:
        pat = os.environ.get("AZDO_PAT") or os.environ.get("AZURE_DEVOPS_EXT_PAT")
        if not pat:
            sys.exit("Sätt AZDO_PAT (Personal Access Token med Work Items Read & Write).")
        token = base64.b64encode(f":{pat}".encode()).decode()
        session = requests.Session()
        session.headers.update({"Authorization": f"Basic {token}"})

    created, failed = 0, 0
    for rec in rows:
        wi_type = map_type(rec.get("Typ"))
        title = str(rec.get("Rubrik") or "").strip()
        fs_id = str(rec.get("ID") or "").strip()
        patch = build_patch(rec)

        if args.dry_run:
            prio = map_priority(rec.get("Prioritering av verksamhet"))
            print(f"[DRY-RUN] {fs_id}: {wi_type} (prio={prio}) — {title}")
            continue

        try:
            wi = create_work_item(session, wi_type, patch)
            wid = wi.get("id")
            print(f"OK  {fs_id}: skapade {wi_type} #{wid} — {title}")
            print(f"    https://dev.azure.com/{ORG}/{PROJECT}/_workitems/edit/{wid}")
            created += 1
        except requests.HTTPError as e:
            body = e.response.text if e.response is not None else str(e)
            print(f"FEL {fs_id}: {title}\n    {body}", file=sys.stderr)
            failed += 1

    if not args.dry_run:
        print(f"\nKlart. Skapade: {created}, Misslyckade: {failed}")


if __name__ == "__main__":
    main()
